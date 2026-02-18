import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.preprocessing import LabelEncoder
import warnings
warnings.filterwarnings("ignore")

np.random.seed(42)

players_raw = [
    {"player_id": 1,  "name": "Virat Kohli",        "role": "BAT",    "country": "India", "matches": 274, "innings": 265, "runs": 12898, "highest": 183, "avg": 57.32, "sr": 139.2, "hundreds": 46, "fifties": 64, "wickets": 4,   "bowling_avg": None, "economy": None,  "catches": 138},
    {"player_id": 2,  "name": "Rohit Sharma",        "role": "BAT",    "country": "India", "matches": 243, "innings": 239, "runs": 9825,  "highest": 264, "avg": 48.60, "sr": 140.1, "hundreds": 30, "fifties": 47, "wickets": 9,   "bowling_avg": None, "economy": None,  "catches": 151},
    {"player_id": 3,  "name": "Shubman Gill",        "role": "BAT",    "country": "India", "matches": 67,  "innings": 64,  "runs": 3124,  "highest": 208, "avg": 51.21, "sr": 104.8, "hundreds": 9,  "fifties": 15, "wickets": 0,   "bowling_avg": None, "economy": None,  "catches": 42},
    {"player_id": 4,  "name": "KL Rahul",            "role": "WK/BAT", "country": "India", "matches": 125, "innings": 117, "runs": 6401,  "highest": 112, "avg": 44.70, "sr": 135.8, "hundreds": 5,  "fifties": 42, "wickets": 0,   "bowling_avg": None, "economy": None,  "catches": 97},
    {"player_id": 5,  "name": "Suryakumar Yadav",    "role": "BAT",    "country": "India", "matches": 67,  "innings": 64,  "runs": 3423,  "highest": 117, "avg": 50.34, "sr": 186.4, "hundreds": 8,  "fifties": 16, "wickets": 0,   "bowling_avg": None, "economy": None,  "catches": 34},
    {"player_id": 6,  "name": "Hardik Pandya",       "role": "ALL",    "country": "India", "matches": 112, "innings": 88,  "runs": 3185,  "highest": 92,  "avg": 32.50, "sr": 148.7, "hundreds": 0,  "fifties": 21, "wickets": 158, "bowling_avg": 27.4, "economy": 8.9,   "catches": 56},
    {"player_id": 7,  "name": "Ravindra Jadeja",     "role": "ALL",    "country": "India", "matches": 198, "innings": 142, "runs": 2756,  "highest": 87,  "avg": 26.40, "sr": 127.3, "hundreds": 0,  "fifties": 12, "wickets": 244, "bowling_avg": 33.1, "economy": 7.6,   "catches": 198},
    {"player_id": 8,  "name": "MS Dhoni",            "role": "WK/BAT", "country": "India", "matches": 350, "innings": 297, "runs": 10773, "highest": 183, "avg": 50.57, "sr": 135.9, "hundreds": 10, "fifties": 73, "wickets": 1,   "bowling_avg": None, "economy": None,  "catches": 321},
    {"player_id": 9,  "name": "Jasprit Bumrah",      "role": "BOWL",   "country": "India", "matches": 148, "innings": 27,  "runs": 56,    "highest": 10,  "avg": 5.20,  "sr": 81.2,  "hundreds": 0,  "fifties": 0,  "wickets": 349, "bowling_avg": 21.7, "economy": 6.6,   "catches": 33},
    {"player_id": 10, "name": "Mohammed Shami",      "role": "BOWL",   "country": "India", "matches": 102, "innings": 18,  "runs": 42,    "highest": 17,  "avg": 4.80,  "sr": 76.4,  "hundreds": 0,  "fifties": 0,  "wickets": 287, "bowling_avg": 23.0, "economy": 7.2,   "catches": 22},
    {"player_id": 11, "name": "Kuldeep Yadav",       "role": "BOWL",   "country": "India", "matches": 98,  "innings": 14,  "runs": 78,    "highest": 25,  "avg": 7.10,  "sr": 92.3,  "hundreds": 0,  "fifties": 0,  "wickets": 198, "bowling_avg": 26.2, "economy": 8.1,   "catches": 12},
    {"player_id": 12, "name": "Yuzvendra Chahal",    "role": "BOWL",   "country": "India", "matches": 78,  "innings": 12,  "runs": 65,    "highest": 18,  "avg": 6.80,  "sr": 88.0,  "hundreds": 0,  "fifties": 0,  "wickets": 205, "bowling_avg": 25.8, "economy": 7.9,   "catches": 10},
    {"player_id": 13, "name": "Mohammed Siraj",      "role": "BOWL",   "country": "India", "matches": 88,  "innings": 20,  "runs": 48,    "highest": 14,  "avg": 5.50,  "sr": 78.9,  "hundreds": 0,  "fifties": 0,  "wickets": 176, "bowling_avg": 25.3, "economy": 7.4,   "catches": 18},
    {"player_id": 14, "name": "Axar Patel",          "role": "ALL",    "country": "India", "matches": 88,  "innings": 58,  "runs": 1124,  "highest": 64,  "avg": 22.48, "sr": 119.5, "hundreds": 0,  "fifties": 4,  "wickets": 142, "bowling_avg": 28.9, "economy": 7.1,   "catches": 44},
    {"player_id": 15, "name": "Shreyas Iyer",        "role": "BAT",    "country": "India", "matches": 84,  "innings": 78,  "runs": 2674,  "highest": 105, "avg": 38.75, "sr": 108.4, "hundreds": 3,  "fifties": 18, "wickets": 0,   "bowling_avg": None, "economy": None,  "catches": 28},
]

opponents = ["Australia", "England", "Pakistan", "New Zealand", "South Africa", "West Indies", "Sri Lanka", "Bangladesh"]
venues    = ["Wankhede", "Eden Gardens", "Chinnaswamy", "Lords", "MCG", "SCG", "Headingley", "Newlands"]
pitches   = ["Flat", "Green", "Dusty", "Bouncy"]

n_matches = 120
match_dates = pd.date_range("2020-01-01", periods=n_matches, freq="9D")
match_results = pd.DataFrame({
    "match_id":       range(1, n_matches + 1),
    "date":           match_dates,
    "format":         np.random.choice(["ODI", "T20I"], n_matches, p=[0.55, 0.45]),
    "venue":          np.random.choice(venues, n_matches),
    "pitch_type":     np.random.choice(pitches, n_matches),
    "opposition":     np.random.choice(opponents, n_matches), 
    "toss_won":       np.random.choice(["Yes", "No"], n_matches),
    "toss_decision":  np.random.choice(["Bat", "Field"], n_matches),
    "india_score":    np.random.randint(145, 340, n_matches),
    "opp_score":      np.random.randint(130, 320, n_matches),
})
match_results["result"]            = np.where(match_results["india_score"] > match_results["opp_score"], "Win", "Loss")
match_results["margin_runs"]       = (match_results["india_score"] - match_results["opp_score"]).abs()
match_results["player_of_match"]   = np.random.choice([p["name"] for p in players_raw], n_matches)
match_results["top_scorer"]        = np.random.choice([p["name"] for p in players_raw[:8]], n_matches)
match_results["top_wicket_taker"]  = np.random.choice([p["name"] for p in players_raw[8:]], n_matches)

le_opp   = LabelEncoder().fit(opponents)
le_pitch = LabelEncoder().fit(pitches)
le_venue = LabelEncoder().fit(venues)

X = pd.DataFrame({
    "opp_enc":        le_opp.transform(match_results["opposition"]),
    "pitch_enc":      le_pitch.transform(match_results["pitch_type"]),
    "venue_enc":      le_venue.transform(match_results["venue"]),
    "toss_won":       (match_results["toss_won"] == "Yes").astype(int),
    "toss_bat":       (match_results["toss_decision"] == "Bat").astype(int),
    "india_avg_sr":   np.full(n_matches, 141.2),
    "bowling_econ":   np.full(n_matches, 7.3),
})
y = (match_results["result"] == "Win").astype(int)

model = GradientBoostingClassifier(n_estimators=80, max_depth=3, random_state=42)
model.fit(X, y)

match_results["win_probability_pct"] = (model.predict_proba(X)[:, 1] * 100).round(1)
match_results["predicted_result"]    = np.where(match_results["win_probability_pct"] >= 50, "Win", "Loss")
match_results["model_correct"]       = (match_results["predicted_result"] == match_results["result"])

upcoming_opponents = opponents
upcoming_fixtures = []
for opp in upcoming_opponents:
    for pitch in pitches:
        row = {
            "fixture":         f"India vs {opp}",
            "opposition":      opp,
            "pitch_type":      pitch,
            "venue":           "TBD",
            "toss_won":        1,
            "toss_bat":        1,
            "india_avg_sr":    141.2,
            "bowling_econ":    7.3,
        }
        row["opp_enc"]   = le_opp.transform([opp])[0]
        row["pitch_enc"] = le_pitch.transform([pitch])[0]
        row["venue_enc"] = 0
        upcoming_fixtures.append(row)

upcoming_df = pd.DataFrame(upcoming_fixtures)
feat_cols   = ["opp_enc", "pitch_enc", "venue_enc", "toss_won", "toss_bat", "india_avg_sr", "bowling_econ"]
upcoming_df["win_probability_pct"] = (model.predict_proba(upcoming_df[feat_cols])[:, 1] * 100).round(1)
upcoming_df["prediction"]          = np.where(upcoming_df["win_probability_pct"] >= 50, "Win", "Loss")
upcoming_df["confidence"]          = pd.cut(upcoming_df["win_probability_pct"],
                                             bins=[0, 40, 55, 70, 100],
                                             labels=["Low", "Medium", "High", "Very High"])
win_predictions = upcoming_df[["fixture", "opposition", "pitch_type", "win_probability_pct", "prediction", "confidence"]].copy()

df_players = pd.DataFrame(players_raw)

def composite_score(row):
    bat_score  = min(row["avg"] * 0.4 + row["sr"] * 0.1 + row["hundreds"] * 2 + row["fifties"], 100)
    bowl_score = (100 - row["bowling_avg"]) * 0.5 + (12 - row["economy"]) * 3 if pd.notna(row["bowling_avg"]) else 0
    field_score = min(row["catches"] * 0.1, 10)
    weights = {"BAT": (0.85, 0.05, 0.10), "WK/BAT": (0.80, 0.05, 0.15),
               "ALL": (0.45, 0.45, 0.10), "BOWL": (0.05, 0.85, 0.10)}
    w = weights.get(row["role"], (0.5, 0.4, 0.1))
    return round(bat_score * w[0] + bowl_score * w[1] + field_score * w[2], 1)

df_players["composite_score"] = df_players.apply(composite_score, axis=1)
df_players["form_score"]      = np.random.uniform(60, 99, len(df_players)).round(1)
df_players["overall_rating"]  = (df_players["composite_score"] * 0.7 + df_players["form_score"] * 0.3).round(1)
df_players_sorted = df_players.sort_values("overall_rating", ascending=False).reset_index(drop=True)

best_xi_ids = []
role_quotas = {"BAT": 4, "WK/BAT": 1, "ALL": 2, "BOWL": 4}
role_counts  = {r: 0 for r in role_quotas}
for _, row in df_players_sorted.iterrows():
    r = row["role"]
    if role_counts.get(r, 0) < role_quotas.get(r, 0):
        best_xi_ids.append(row["player_id"])
        role_counts[r] = role_counts.get(r, 0) + 1
    if len(best_xi_ids) == 11:
        break

best_xi = df_players[df_players["player_id"].isin(best_xi_ids)].copy()
best_xi["xi_position"] = range(1, len(best_xi) + 1)
best_xi = best_xi[["xi_position", "name", "role", "avg", "sr", "wickets", "economy", "catches", "composite_score", "form_score", "overall_rating"]]

innings_rows = []
for p in players_raw[:11]:
    scores = np.random.randint(0, 140, 8).tolist() if p["role"] != "BOWL" else [0] * 8
    wickets_per = np.random.randint(0, 5, 8).tolist() if p["role"] in ("BOWL", "ALL") else [0] * 8
    for i in range(8):
        innings_rows.append({
            "player_id":  p["player_id"],
            "name":       p["name"],
            "role":       p["role"],
            "innings_no": i + 1,
            "runs":       scores[i],
            "wickets":    wickets_per[i],
            "strike_rate": round(scores[i] / max(np.random.randint(20, 100), 1) * 100, 1),
        })
trends = pd.DataFrame(innings_rows)
trends["rolling_avg_runs"]    = trends.groupby("player_id")["runs"].transform(lambda x: x.expanding().mean().round(1))
trends["rolling_avg_wickets"] = trends.groupby("player_id")["wickets"].transform(lambda x: x.expanding().mean().round(2))

output_path = r"C:\Users\bajpa\OneDrive\Desktop\dashboard\cricket_analytics_tableau.xlsx"

sheet_data = {
    "player_stats":        df_players.drop(columns=["composite_score"]),
    "match_results":       match_results,
    "win_predictions":     win_predictions,
    "best_xi":             best_xi,
    "performance_trends":  trends,
}

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    for sheet_name, df in sheet_data.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)

wb = load_workbook(output_path)

COLORS = {
    "header_fill":  "1A1F36",
    "header_font":  "F0B429",
    "accent_green": "56CFAD",
    "accent_red":   "E05C5C",
    "alt_row":      "F4F6FA",
    "bat_role":     "FFF3CD",
    "bowl_role":    "FFDEDE",
    "all_role":     "D9E6FF",
    "wk_role":      "D6F5EB",
}

header_font    = Font(name="Arial", bold=True, color=COLORS["header_font"], size=11)
header_fill    = PatternFill("solid", start_color=COLORS["header_fill"])
header_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
center_align   = Alignment(horizontal="center", vertical="center")
thin_border    = Border(
    bottom=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="EEEEEE"),
)

role_fills = {
    "BAT":    PatternFill("solid", start_color=COLORS["bat_role"]),
    "BOWL":   PatternFill("solid", start_color=COLORS["bowl_role"]),
    "ALL":    PatternFill("solid", start_color=COLORS["all_role"]),
    "WK/BAT": PatternFill("solid", start_color=COLORS["wk_role"]),
}

def style_sheet(ws, role_col=None, highlight_col=None, highlight_thresh=None, highlight_color="56CFAD", highlight_low_color="E05C5C"):
    # Header row
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
    ws.row_dimensions[1].height = 32

    # Data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        alt_fill = PatternFill("solid", start_color=COLORS["alt_row"]) if row_idx % 2 == 0 else None
        for cell in row:
            cell.border    = thin_border
            cell.alignment = center_align
            if alt_fill:
                cell.fill = alt_fill

        # Role-based row colouring
        if role_col:
            role_cell = ws.cell(row=row_idx, column=role_col)
            role_val  = role_cell.value
            if role_val in role_fills:
                for cell in row:
                    cell.fill = role_fills[role_val]

        # Win-probability highlight
        if highlight_col:
            hc = ws.cell(row=row_idx, column=highlight_col)
            try:
                v = float(hc.value)
                if v >= highlight_thresh:
                    hc.fill = PatternFill("solid", start_color=highlight_color)
                    hc.font = Font(name="Arial", bold=True, color="FFFFFF")
                elif v < 45:
                    hc.fill = PatternFill("solid", start_color=highlight_low_color)
                    hc.font = Font(name="Arial", bold=True, color="FFFFFF")
            except (TypeError, ValueError):
                pass

    # Auto column width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 28)

    # Freeze top row
    ws.freeze_panes = "A2"

# Apply styling
role_col_idx = {
    "player_stats":    list(df_players.columns).index("role") + 1,
    "best_xi":         list(best_xi.columns).index("role") + 1,
    "performance_trends": list(trends.columns).index("role") + 1,
}

win_prob_col = list(win_predictions.columns).index("win_probability_pct") + 1
result_col   = list(match_results.columns).index("win_probability_pct") + 1

style_sheet(wb["player_stats"],        role_col=role_col_idx["player_stats"])
style_sheet(wb["match_results"],       highlight_col=result_col, highlight_thresh=60)
style_sheet(wb["win_predictions"],     highlight_col=win_prob_col, highlight_thresh=60)
style_sheet(wb["best_xi"],             role_col=role_col_idx["best_xi"])
style_sheet(wb["performance_trends"],  role_col=role_col_idx["performance_trends"])

readme = wb.create_sheet(" Tableau Guide")
guide_rows = [
    ["CricIQ Analytics — Tableau Connection Guide", "", ""],
    ["", "", ""],
    ["Sheet", "Tableau Chart Type", "Suggested Fields"],
    ["player_stats",        "Scatter / Bar",        "avg vs sr (size=runs), color by role"],
    ["match_results",       "Line / Calendar",      "date vs india_score, color by result"],
    ["match_results",       "Bar",                  "opposition vs win_probability_pct"],
    ["win_predictions",     "Heatmap",              "opposition × pitch_type, color by win_probability_pct"],
    ["win_predictions",     "Bullet / Gauge",       "win_probability_pct, color by confidence"],
    ["best_xi",             "Lollipop / Bar",       "name vs overall_rating, color by role"],
    ["performance_trends",  "Line (multi-series)",  "innings_no vs rolling_avg_runs, filter by name"],
    ["performance_trends",  "Bump chart",           "innings_no vs rolling_avg_wickets for bowlers"],
    ["", "", ""],
    ["Model Accuracy", f"{(match_results['model_correct'].mean()*100):.1f}%", "GradientBoostingClassifier, 80 estimators"],
    ["Total Matches",  str(n_matches), "Jan 2020 – present"],
    ["Players Tracked", str(len(players_raw)), "India squad (ODI + T20I)"],
]
for r_idx, row_data in enumerate(guide_rows, 1):
    for c_idx, val in enumerate(row_data, 1):
        cell = readme.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == 1:
            cell.font = Font(name="Arial", bold=True, color=COLORS["header_font"], size=14)
            cell.fill = PatternFill("solid", start_color=COLORS["header_fill"])
        elif r_idx == 3:
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill("solid", start_color="2D3E5A")
        elif r_idx > 3 and val:
            cell.font = Font(name="Arial", size=10)
            if r_idx % 2 == 0:
                cell.fill = PatternFill("solid", start_color=COLORS["alt_row"])
        cell.alignment = Alignment(horizontal="left", vertical="center")

readme.column_dimensions["A"].width = 26
readme.column_dimensions["B"].width = 30
readme.column_dimensions["C"].width = 48
readme.row_dimensions[1].height = 28
readme.freeze_panes = "A4"


wb.move_sheet(" Tableau Guide", offset=-(len(wb.sheetnames)-1))

wb.save(output_path)
print(f" Saved → {output_path}")
print(f" Sheets: {wb.sheetnames}")
print(f" Model accuracy: {(match_results['model_correct'].mean()*100):.1f}%")
print(f" Win predictions generated: {len(win_predictions)} fixture/pitch combos")
